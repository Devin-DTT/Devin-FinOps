"""
Export module for FinOps summary metrics to Excel format.
Provides functionality to export business summary data to Excel with professional formatting.

Uses dataclasses for parameter grouping, a Builder pattern for report construction,
and separates formatting logic from content generation.
"""

import logging
import json
from dataclasses import dataclass, field
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

TARGET_METRICS = [
    'ACUs Mes',
    'ACUs Mes Anterior',
    'Coste Mes',
    'Coste Mes Anterior',
    'Diferencia (Mes Actual - Mes Anterior)',
    'Variacion Porcentual',
    'ACUs consumidos totales',
    'Coste consumo total',
    'Numero de usuarios',
    'Numero de sesiones',
    'Media total ACUs por usuario',
    'Media total Coste por usuario',
    'Media mes ACUs por usuario',
    'Media mes Coste por usuario',
    'Media Total ACUs por sesion',
    'Media Total Coste por sesion',
    'Numero de organizaciones',
    'Coste por organizacion',
    'Coste por grupo (IdP)',
    'Numero de PRs abiertos',
    'Numero de PRs cerradas',
    'Numero de PRs mergeadas',
    'Numero de PRs totales',
    'ACUs por PR totales',
    'Coste por PR totales',
    'ACUs medio por PR mergeado',
    'Coste medio por PR mergeado',
    'Promedio de PRs por ACU',
    'Promedio de PRs mergeados por ACU',
    'Tasa de exito de PR'
]


@dataclass
class ExportConfig:
    """Configuration for Excel export output."""
    output_filename: str = 'finops_summary_report.xlsx'
    config: Any = None


@dataclass
class ReportData:
    """Data required for the main FinOps report sheet."""
    consumption_data: Dict[str, Any] = field(default_factory=dict)
    all_api_data: Optional[Dict[str, Dict[str, Any]]] = None
    summary_data: Optional[Dict[str, Any]] = None
    finops_metrics: Optional[Dict[str, Dict[str, Any]]] = None


@dataclass
class BreakdownData:
    """Data for breakdown and history sheets."""
    user_breakdown_list: Optional[List[Dict[str, Any]]] = None
    org_breakdown_summary: Optional[Dict[str, Any]] = None
    monthly_consumption_history: Optional[List[Dict[str, Any]]] = None


class ExcelFormatter:
    """Encapsulates all Excel styling and formatting concerns."""

    def __init__(self) -> None:
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.section_font = Font(bold=True, size=13)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def apply_section_title(self, ws: Worksheet, row: int, title: str, merge_end_col: str = 'E') -> None:
        cell_ref = f'A{row}'
        ws[cell_ref] = title
        ws[cell_ref].font = self.section_font
        ws[cell_ref].fill = self.section_fill
        ws.merge_cells(f'A{row}:{merge_end_col}{row}')

    def apply_header_row(self, ws: Worksheet, row: int, headers: List[str]) -> None:
        columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        for i, header in enumerate(headers):
            cell_ref = f'{columns[i]}{row}'
            ws[cell_ref] = header
            ws[cell_ref].font = self.header_font
            ws[cell_ref].fill = self.header_fill
            ws[cell_ref].alignment = self.header_alignment

    def apply_column_widths(self, ws: Worksheet, widths: Dict[str, int]) -> None:
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    def format_numeric_cell(self, ws: Worksheet, cell_ref: str, value: Any, number_format: str = '0.00') -> None:
        if isinstance(value, (int, float)) and value != 'N/A':
            ws[cell_ref] = round(value, 2)
            ws[cell_ref].number_format = number_format
        else:
            ws[cell_ref] = str(value)

    def apply_wrap_top_alignment(self, ws: Worksheet, cell_ref: str) -> None:
        ws[cell_ref].alignment = Alignment(wrap_text=True, vertical="top")

    def apply_right_top_alignment(self, ws: Worksheet, cell_ref: str) -> None:
        ws[cell_ref].alignment = Alignment(horizontal="right", vertical="top")

    def apply_right_alignment(self, ws: Worksheet, cell_ref: str) -> None:
        ws[cell_ref].alignment = Alignment(horizontal="right")


def _extract_consumption_field(all_api_data: Optional[Dict[str, Dict[str, Any]]], field_name: str) -> dict:
    """Extract a field from the consumption_daily API response."""
    if not all_api_data or not isinstance(all_api_data, dict):
        return {}
    consumption_endpoint = all_api_data.get('consumption_daily', {})
    if consumption_endpoint.get('status_code') != 200:
        return {}
    try:
        response_data = consumption_endpoint.get('response', {})
        if isinstance(response_data, str):
            response_data = json.loads(response_data)
        if isinstance(response_data, dict):
            return response_data.get(field_name, {})
    except (json.JSONDecodeError, AttributeError, TypeError) as e:
        logger.warning(f"Failed to extract {field_name}: {e}")
    return {}


def _generate_finops_sheet_content(ws: Worksheet, finops_metrics: Optional[Dict[str, Dict[str, Any]]], fmt: ExcelFormatter) -> None:
    """Generate the FinOps Report sheet content and apply formatting."""
    fmt.apply_column_widths(ws, {'A': 35, 'B': 20, 'C': 50, 'D': 50, 'E': 25})

    row = 1
    fmt.apply_section_title(ws, row, "Visibilidad y transparencia de costes")
    row += 1

    fmt.apply_header_row(ws, row, ["METRICA FINOPS", "RESULTADO", "FORMULA / LOGICA", "FUENTE JSON", "DATO EXTRAÍDO"])
    row += 1

    if not finops_metrics:
        logger.warning("No finops_metrics provided, cannot populate Excel report")
        return

    for metric_name in TARGET_METRICS:
        if metric_name not in finops_metrics:
            continue
        metric_data = finops_metrics[metric_name]

        ws[f'A{row}'] = metric_name
        fmt.apply_wrap_top_alignment(ws, f'A{row}')

        value = metric_data.get('value', 'N/A')
        fmt.format_numeric_cell(ws, f'B{row}', value)
        fmt.apply_right_top_alignment(ws, f'B{row}')

        if 'formula' in metric_data:
            ws[f'C{row}'] = metric_data.get('formula', '')
            fmt.apply_wrap_top_alignment(ws, f'C{row}')

            sources_used = metric_data.get('sources_used', [])
            if sources_used:
                source_paths = [source.get('source_path', '') for source in sources_used]
                ws[f'D{row}'] = '\n'.join(source_paths)
            else:
                ws[f'D{row}'] = ''
            fmt.apply_wrap_top_alignment(ws, f'D{row}')

            raw_data = metric_data.get('raw_data_value', 'N/A')
            fmt.format_numeric_cell(ws, f'E{row}', raw_data)
            fmt.apply_right_top_alignment(ws, f'E{row}')

        elif 'external_data_required' in metric_data:
            ws[f'C{row}'] = metric_data.get('reason', '')
            fmt.apply_wrap_top_alignment(ws, f'C{row}')
            ws[f'D{row}'] = metric_data.get('external_data_required', '')
            fmt.apply_wrap_top_alignment(ws, f'D{row}')
            ws[f'E{row}'] = 'N/A'
            fmt.apply_right_top_alignment(ws, f'E{row}')

        row += 1

    logger.info("Added 'Visibilidad y transparencia de costes' section with metrics")


def _generate_user_sheet_content(wb: Workbook, consumption_by_user: dict, fmt: ExcelFormatter) -> None:
    """Generate the User Breakdown sheet."""
    if not consumption_by_user:
        logger.warning("No consumption_by_user data available for 'Desglose Consumo Usuario' sheet")
        return

    ws_user = wb.create_sheet(title="Desglose Consumo Usuario")
    fmt.apply_column_widths(ws_user, {'A': 40, 'B': 20})
    fmt.apply_header_row(ws_user, 1, ["User ID", "Consumo (ACUs)"])

    row = 2
    for user_id, acus in consumption_by_user.items():
        ws_user[f'A{row}'] = user_id
        ws_user[f'B{row}'] = round(acus, 2)
        ws_user[f'B{row}'].number_format = '0.00'
        fmt.apply_right_alignment(ws_user, f'B{row}')
        row += 1

    logger.info(f"Created 'Desglose Consumo Usuario' sheet with {len(consumption_by_user)} users")


def _generate_org_sheet_content(wb: Workbook, consumption_by_org_id: dict, fmt: ExcelFormatter) -> None:
    """Generate the Organization Breakdown sheet."""
    if not consumption_by_org_id:
        logger.warning("No consumption_by_org_id data available for 'Desglose Consumo Organizacion' sheet")
        return

    ws_org = wb.create_sheet(title="Desglose Consumo Organizacion")
    fmt.apply_column_widths(ws_org, {'A': 40, 'B': 20})
    fmt.apply_header_row(ws_org, 1, ["Organization ID", "Consumo (ACUs)"])

    row = 2
    for org_id, acus in consumption_by_org_id.items():
        ws_org[f'A{row}'] = org_id
        ws_org[f'B{row}'] = round(acus, 2)
        ws_org[f'B{row}'].number_format = '0.00'
        fmt.apply_right_alignment(ws_org, f'B{row}')
        row += 1

    logger.info(f"Created 'Desglose Consumo Organizacion' sheet with {len(consumption_by_org_id)} organizations")


def _generate_monthly_sheet_content(wb: Workbook, monthly_consumption_history: Optional[List[Dict[str, Any]]], fmt: ExcelFormatter) -> None:
    """Generate the Monthly History sheet."""
    if not monthly_consumption_history:
        logger.warning("No monthly_consumption_history data available for 'Historico Consumo Mensual' sheet")
        return

    ws_monthly = wb.create_sheet(title="Historico Consumo Mensual")
    fmt.apply_column_widths(ws_monthly, {'A': 20, 'B': 25})
    fmt.apply_header_row(ws_monthly, 1, ["Mes", "Consumo Total (ACUs)"])

    row = 2
    for month_data in monthly_consumption_history:
        mes = month_data.get('Mes', '')
        acus = month_data.get('Consumo Total (ACUs)', 0)
        ws_monthly[f'A{row}'] = mes
        ws_monthly[f'B{row}'] = acus
        ws_monthly[f'B{row}'].number_format = '0.00'
        fmt.apply_right_alignment(ws_monthly, f'B{row}')
        row += 1

    logger.info(f"Created 'Historico Consumo Mensual' sheet with {len(monthly_consumption_history)} months")


class ReportBuilder:
    """Builder pattern for constructing FinOps Excel reports step by step."""

    def __init__(self, export_config: Optional[ExportConfig] = None) -> None:
        self._export_config = export_config or ExportConfig()
        self._report_data: Optional[ReportData] = None
        self._breakdown_data: Optional[BreakdownData] = None
        self._formatter = ExcelFormatter()
        self._wb = Workbook()
        self._finops_sheet_added = False
        self._user_sheet_added = False
        self._org_sheet_added = False
        self._monthly_sheet_added = False

    def with_report_data(self, report_data: ReportData) -> 'ReportBuilder':
        self._report_data = report_data
        return self

    def with_breakdown_data(self, breakdown_data: BreakdownData) -> 'ReportBuilder':
        self._breakdown_data = breakdown_data
        return self

    def with_finops_metrics(self, finops_metrics: Dict[str, Dict[str, Any]]) -> 'ReportBuilder':
        if self._report_data is None:
            self._report_data = ReportData()
        self._report_data.finops_metrics = finops_metrics
        return self

    def with_user_breakdown(self, user_breakdown_list: List[Dict[str, Any]]) -> 'ReportBuilder':
        if self._breakdown_data is None:
            self._breakdown_data = BreakdownData()
        self._breakdown_data.user_breakdown_list = user_breakdown_list
        return self

    def with_org_breakdown(self, org_breakdown_summary: Dict[str, Any]) -> 'ReportBuilder':
        if self._breakdown_data is None:
            self._breakdown_data = BreakdownData()
        self._breakdown_data.org_breakdown_summary = org_breakdown_summary
        return self

    def with_monthly_history(self, monthly_consumption_history: List[Dict[str, Any]]) -> 'ReportBuilder':
        if self._breakdown_data is None:
            self._breakdown_data = BreakdownData()
        self._breakdown_data.monthly_consumption_history = monthly_consumption_history
        return self

    def add_finops_sheet(self) -> 'ReportBuilder':
        finops_metrics = self._report_data.finops_metrics if self._report_data else None
        ws = self._wb.active
        ws.title = "FinOps Report"
        _generate_finops_sheet_content(ws, finops_metrics, self._formatter)
        self._finops_sheet_added = True
        return self

    def add_user_sheet(self) -> 'ReportBuilder':
        all_api_data = self._report_data.all_api_data if self._report_data else None
        consumption_by_user = _extract_consumption_field(all_api_data, 'consumption_by_user')
        _generate_user_sheet_content(self._wb, consumption_by_user, self._formatter)
        self._user_sheet_added = True
        return self

    def add_org_sheet(self) -> 'ReportBuilder':
        all_api_data = self._report_data.all_api_data if self._report_data else None
        consumption_by_org_id = _extract_consumption_field(all_api_data, 'consumption_by_org_id')
        _generate_org_sheet_content(self._wb, consumption_by_org_id, self._formatter)
        self._org_sheet_added = True
        return self

    def add_monthly_history_sheet(self) -> 'ReportBuilder':
        history = self._breakdown_data.monthly_consumption_history if self._breakdown_data else None
        _generate_monthly_sheet_content(self._wb, history, self._formatter)
        self._monthly_sheet_added = True
        return self

    def build(self) -> None:
        output = self._export_config.output_filename
        logger.info(f"Exporting summary data to {output}...")

        try:
            if not self._finops_sheet_added:
                self.add_finops_sheet()
            if not self._user_sheet_added:
                self.add_user_sheet()
            if not self._org_sheet_added:
                self.add_org_sheet()
            if not self._monthly_sheet_added:
                self.add_monthly_history_sheet()

            self._wb.save(output)
            logger.info(f"Successfully exported summary to {output}")
            print(f"\n+ Summary data exported to {output}")

        except Exception as e:
            logger.error(f"Failed to export summary to Excel: {e}", exc_info=True)
            print(f"\n- Error exporting summary to Excel: {e}")


def export_summary_to_excel(
    consumption_data: Dict[str, Any],
    config: Any,
    all_api_data: Dict[str, Dict[str, Any]] = None,
    output_filename: str = 'finops_summary_report.xlsx',
    summary_data: Dict[str, Any] = None,
    user_breakdown_list: list = None,
    org_breakdown_summary: dict = None,
    finops_metrics: Dict[str, Dict[str, Any]] = None,
    monthly_consumption_history: list = None
) -> None:
    """
    Export FinOps business summary data to Excel format with modular structure.
    Backward-compatible wrapper that delegates to ReportBuilder.

    Args:
        consumption_data: Dictionary containing metrics and reporting_period
        config: MetricsConfig object with pricing information
        all_api_data: Raw API response data for all endpoints
        output_filename: Output Excel filename
        summary_data: Generated business summary data
        user_breakdown_list: List of user consumption breakdown dictionaries
        org_breakdown_summary: Dictionary of organization aggregation data
        finops_metrics: Dictionary containing FinOps metrics with traceability information
        monthly_consumption_history: List of monthly ACU consumption records
    """
    export_config = ExportConfig(output_filename=output_filename, config=config)
    report_data = ReportData(
        consumption_data=consumption_data or {},
        all_api_data=all_api_data,
        summary_data=summary_data,
        finops_metrics=finops_metrics,
    )
    breakdown_data = BreakdownData(
        user_breakdown_list=user_breakdown_list,
        org_breakdown_summary=org_breakdown_summary,
        monthly_consumption_history=monthly_consumption_history,
    )

    builder = ReportBuilder(export_config)
    builder.with_report_data(report_data)
    builder.with_breakdown_data(breakdown_data)
    builder.build()
